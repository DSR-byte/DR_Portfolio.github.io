import tkinter as tk
from tkinter import filedialog, scrolledtext
from threading import Thread
from Lead_Scraper_Execel_Feader import ExcelFeeder
import sys, io, time
from openpyxl import Workbook


class TextRedirector(io.StringIO):
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget

    def write(self, s):
        try:
            if self.text_widget.winfo_exists() and self.text_widget.winfo_ismapped():
                self.text_widget.after(0, self._safe_insert, s)
        except (tk.TclError, AttributeError):
            pass

    def _safe_insert(self, s):
        self.text_widget.insert(tk.END, s)
        self.text_widget.see(tk.END)

    def flush(self):
        pass


class Lead_Scraper_UI:
    def __init__(self, root):
        print("Initializing GUI application...")

        self.root, self.filename, self.feeder, self.scraper_thread = root, None, None, None
        self.current_row_index = 0
        self.results = []

        root.title("Lead Scraper DSR1")

        main = tk.Frame(root)
        main.grid(row=0, column=0, sticky='nsew')
        root.grid_rowconfigure(0, weight=1)
        root.grid_columnconfigure(0, weight=1)

        tk.Button(main, text="Select Excel File", command=self.load_file).grid(row=0, column=0, sticky='w', padx=10, pady=10)
        tk.Button(main, text="Start", command=self.start_scraper).grid(row=8, column=6, padx=10)
        tk.Button(main, text="Pause", command=self.pause_scraper).grid(row=8, column=7, padx=10)
        tk.Button(main, text="Resume", command=self.resume_scraper).grid(row=8, column=8, padx=10)
        tk.Button(main, text="Save Results", command=self.save_results).grid(row=8, column=9, padx=10)

        self.status = tk.Label(main, text="No file selected.")
        self.status.grid(row=1, column=0, sticky='w', padx=10)

        self.output_box = scrolledtext.ScrolledText(main, width=40, height=10)
        self.output_box.grid(row=3, column=0, columnspan=2, sticky='sew', padx=10, pady=10)

        self.start_label = tk.Label(main, text="Start From Row:")
        self.start_label.grid(row=5, column=0, sticky='w', padx=10, pady=(0, 10))
        self.start_entry = tk.Entry(main, width=10)
        self.start_entry.insert(0, "0")
        self.start_entry.grid(row=5, column=1, sticky='w', padx=10, pady=(0, 10))

        self.current_display = tk.Text(
            main,
            font=("Helvetica", 18, "bold"),
            bg="white",
            fg="black",
            relief="sunken",
            width=60,
            height=30
        )
        self.current_display.grid(row=1, column=4, columnspan=2, rowspan=6, sticky='nsew', padx=10, pady=10)
        self.current_display.insert(tk.END, "Current Result:\n—\n")
        self.current_display.config(state='disabled')

        main.grid_rowconfigure(2, weight=1)
        main.grid_columnconfigure(1, weight=1)

        root.protocol("WM_DELETE_WINDOW", self.on_close)

        sys.stdout = TextRedirector(self.output_box)
        sys.stderr = TextRedirector(self.output_box)

        print("GUI initialized successfully.")

    def append_text(self, message):
        self.current_display.config(state='normal')
        self.current_display.insert('end', message + '\n')
        self.current_display.see('end')
        self.current_display.config(state='disabled')

    def start_scraper(self):
        if not self.filename:
            print("Please select a file first.")
            return
        try:
            self.current_row_index = int(self.start_entry.get())
        except ValueError:
            self.current_row_index = 0
            print("⚠️ Invalid row number — defaulting to 0")

        print("Starting scraper...")
        self.stopped = False
        self.paused = False
        self.scraper_thread = Thread(target=self.run_feeder_scraper)
        self.scraper_thread.start()

    def pause_scraper(self):
        print("Pausing...")
        self.feeder.pause_flag = True
        self.paused = True

    def resume_scraper(self):
        print("Resuming...")
        self.paused = False
        self.feeder.pause_flag = False

    def load_file(self):
        print("Opening file dialog...")
        self.filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.status.config(text="File Loaded")

    def run_feeder_scraper(self):
        print("Running feeder scraper...")
        self.feeder = ExcelFeeder(self.filename)
        display_text = ""

        all_rows = list(self.feeder.ws.iter_rows(min_row=self.feeder.start_row, values_only=True))

        for i, (name, zip_code, phone) in enumerate(self.feeder.run_scraper(start_index=self.current_row_index), start=self.current_row_index):
            while self.paused:
                time.sleep(0.3)
            row = all_rows[i]
            if not hasattr(self.feeder, 'address_idx'):
                headers = [str(cell).strip() for cell in next(self.feeder.ws.iter_rows(min_row=2, max_row=2, values_only=True))]
                self.feeder.address_idx = headers.index("Site Address")
                self.feeder.city_idx = headers.index("Site City")
                self.feeder.state_idx = headers.index("Site State")

            site_address = row[self.feeder.address_idx]
            site_city = row[self.feeder.city_idx]
            site_state = row[self.feeder.state_idx]
            full_address = f"{site_address}, {site_city}, {site_state}, {zip_code}"

            line = f"📱 {phone:<15} 👤 {name:<25} 🏠 {full_address}\n" + "—" * 100 + "\n"
            display_text += line

            self.current_display.config(state='normal')
            self.current_display.delete("1.0", tk.END)
            self.current_display.insert(tk.END, display_text)
            self.current_display.see(tk.END)
            self.current_display.config(state='disabled')

            self.output_box.insert(tk.END, line)
            self.output_box.see(tk.END)

            self.results.append((name, full_address, phone))

        print("✅ Scraping complete.")

    def save_results(self):
        if not self.results:
            print("⚠️ No results to save.")
            return
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Address", "Phone"])
        for row in self.results:
            ws.append(row)
        filename = f"scraping_results_{int(time.time())}.xlsx"
        wb.save(filename)
        print(f"✅ Results saved to {filename}")

    def on_close(self):
        print("Closing app...")
        if self.feeder:
            self.feeder.stop_flag = True
            self.resume_scraper()
        sys.stdout = sys.__stdout__
        sys.stderr = sys.__stderr__
        self.root.quit()
        self.root.destroy()


if __name__ == '__main__':
    print("Launching application...")
    root = tk.Tk()
    app = Lead_Scraper_UI(root)
    root.mainloop()
