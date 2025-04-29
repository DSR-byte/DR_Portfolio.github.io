from openpyxl import load_workbook
from LEAD_SCRAPER_SCRAPER import TruePeopleSearchScraper
import time

class ExcelFeeder:
    def __init__(self, filename):
        self.wb = load_workbook(filename)
        self.ws = self.wb.active
        self.scraper = TruePeopleSearchScraper()
        self.stop_flag = False
        self.pause_flag = False

        # Automatically detect header row
        header_row = None
        for i, row in enumerate(self.ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if any(cell and isinstance(cell, str) and "owner" in cell.lower() for cell in row):
                header_row = i
                self.headers = [str(cell).strip() if cell else "" for cell in row]
                break

        if header_row is None:
            raise ValueError("❌ Could not find a valid header row.")

        print(f"✅ Found header row at line {header_row}:")
        print(self.headers)

        try:
            self.first_idx = self.headers.index("1st Owner's First Name")
            self.last_idx = self.headers.index("1st Owner's Last Name")
            self.zip_idx = self.headers.index("Site Zip Code")
        except ValueError as e:
            raise ValueError("❌ One or more expected column names not found.") from e

        self.start_row = header_row + 1

    def run_scraper(self, start_index=0):
        for i, row in enumerate(self.ws.iter_rows(min_row=self.start_row, values_only=True)):
            if self.stop_flag:
                self.close()
                break
    
            if i < start_index:
                continue
    
            while self.pause_flag:
                time.sleep(0.3)
    
            first = row[self.first_idx]
            last = row[self.last_idx]
            zip_code = row[self.zip_idx]
    
            if not first or not last or not zip_code:
                print("⏭️ Skipping empty or invalid row")
                continue
    
            name = f"{first} {last}"
            location = f"{zip_code} CA"
            print(f"\n📡 Searching for: {name} | {location}")
    
            phone = "N/A"
            for attempt in range(1):
                while self.pause_flag:
                    time.sleep(0.3)
                phone = self.scraper.search(name, location)
                if phone and phone != "N/A":
                    break
                print(f"🔁 Retry {attempt} for {name}...")
                time.sleep(5 + attempt * 2)
    
            print(f"📞 Found: {phone}")
            yield (name, zip_code, phone)
            time.sleep(1)


    def stop(self):
        self.stop_flag = True

    def close(self):
        self.scraper.close()



