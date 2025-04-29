#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Mar  2 22:55:33 2025

@author: davidrock
"""

"this will text lots of people"
"Found at:  /Users/davidrock"
"Name of File: Auto_Texter.py"

import sys
import re
import subprocess
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class SMSSenderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Bulk SMS Sender")
        self.root.geometry("500x300")

        self.file_label = tk.Label(root, text="Select Excel File:")
        self.file_label.pack()
        self.file_entry = tk.Entry(root, width=40)
        self.file_entry.pack()
        self.file_button = tk.Button(root, text="Browse", command=self.browse_file)
        self.file_button.pack()

        self.msg_label = tk.Label(root, text="Enter Message (use {name} and {address} for personalization):")
        self.msg_label.pack()
        self.msg_entry = tk.Text(root, height=4, width=40)
        self.msg_entry.pack()

        self.send_button = tk.Button(root, text="Send Messages", command=self.send_messages)
        self.send_button.pack(pady=5)
        self.quit_button = tk.Button(root, text="Quit", command=self.force_quit)
        self.quit_button.pack(pady=5)

    def force_quit(self):
        self.root.destroy()
        sys.exit(0)

    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.file_entry.delete(0, tk.END)
        self.file_entry.insert(0, file_path)

    def clean_phone_number(self, number):
        number = str(number).replace("\u202a", "").replace("\u202c", "")
        number = re.sub(r"[^\d+]", "", number)

        if number.startswith("+1"):
            return number
        elif number.startswith("1") and len(number) == 11:
            return f"+{number}"
        elif len(number) == 10:
            return f"+1{number}"
        else:
            print(f"Invalid phone number format: {number}")
            return None

    def send_messages(self):
        file_path = self.file_entry.get()
        message_template = self.msg_entry.get("1.0", tk.END).strip()

        if not file_path or not message_template:
            messagebox.showerror("Error", "Please select a file and enter a message.")
            return

        try:
            wb = load_workbook(get_resource_path(file_path))
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            if not {"Names", "Numbers", "Address"}.issubset(headers):
                messagebox.showerror("Error", "Excel file must have 'Names', 'Numbers', and 'Address' columns.")
                return

            name_idx = headers.index("Names")
            number_idx = headers.index("Numbers")
            address_idx = headers.index("Address")

            for row in ws.iter_rows(min_row=2, values_only=True):
                name = row[name_idx]
                number = self.clean_phone_number(row[number_idx])
                address = row[address_idx]

                if not number:
                    continue

                message = message_template.replace("{name}", str(name)).replace("{address}", str(address))
                self.send_sms(name, number, message)

            messagebox.showinfo("Success", "All messages have been sent!")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Excel file: {e}")

    def send_sms(self, name, number, message):
        applescript = f'''
        tell application "Messages"
            set targetBuddy to "{number}"
            set targetService to 1st service whose service type = SMS
            send "{message}" to buddy targetBuddy of targetService
        end tell
        '''
        process = subprocess.run(["osascript", "-e", applescript], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        error = process.stderr.decode()

        if error:
            print(f"Error sending SMS to {name} ({number}): {error}")
        else:
            print(f"SMS sent to {name} ({number})!")

if __name__ == "__main__":
    root = tk.Tk()
    app = SMSSenderApp(root)
    root.mainloop()
